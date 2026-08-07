from app.main import _filter_rows, _columns, _build_workbook
from datetime import datetime
from openpyxl import load_workbook
import io

rows = [
    {"codObjeto":"65", "detObjeto":"Servicio", "fechaConvocatoria":"05/08/2026 09:00", "detEntidad":"A"},
    {"codObjeto":"65", "detObjeto":"Servicio", "fechaConvocatoria":"06/08/2026 10:00", "detEntidad":"B"},
    {"codObjeto":"64", "detObjeto":"Obra", "fechaConvocatoria":"05/08/2026 11:00", "detEntidad":"C"},
]
filtered = _filter_rows(rows, "65", {datetime(2026,8,5)}, None, None)
assert len(filtered) == 1 and filtered[0]["detEntidad"] == "A"
cols = _columns(filtered)
content = _build_workbook(filtered, cols, [("Objeto","Servicio (65)"),("Registros filtrados",1)])
wb = load_workbook(io.BytesIO(content))
assert wb.sheetnames == ["Oportunidades", "Control"]
assert wb["Oportunidades"].max_row == 2
print("OK")
